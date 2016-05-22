import openpyxl
import warnings


class ExtractInfo(object):
    ''' This module allows one to scrape specific information from the 
        All_Yale_&_UK_Variants.xlsx spreadsheet and add all query information for a 
        sepcified query in a dictionary.
    ''' 

    def __init__(self,xlsx,sheet,row_header):

        self.xlsx = xlsx
        self.sheet = sheet
        self.row_header = row_header
    

    # dictionary conating headers as keys
    header_contents = {}


    def open_spreadsheets(self):
        '''Open the given workbook and sheet
        '''   
        # stop warning from printing to the screen
        warnings.simplefilter("ignore")

        # open workbook and sheets (sheet name or number) within
        wb = openpyxl.load_workbook(self.xlsx,data_only=True)
        if self.sheet.isdigit():
            spreadsheet = wb.worksheets[int(self.sheet)]
        else:
            spreadsheet = wb.get_sheet_by_name(self.sheet)   
        
        return spreadsheet
              
              
    def create_header_dicts(self,spreadsheet):
        '''Store each entry in the spreadsheets header as
           a key with no value in the headers_contents dict
        '''
        for column_number in range(1,100):
            header = spreadsheet.cell(row=self.row_header,column=column_number).value
            if header is None:
                continue
            ExtractInfo.header_contents[header] = ''

                             

    def get_row(self,spreadsheet, query):
        ''' Search the database/sheet and match with the query/variant_alias
            if found, output the querys row number in the spreadsheet.
        '''
        counter = 0
        
        for row_number in range(1,500):
            if query == spreadsheet.cell(row=row_number,column=1).value:
                counter += 1
                return row_number
        
        if counter == 0:
            error_message = " ".join((query,"not found in",self.xlsx))
            print(error_message)
    
    
    def get_query_info(self,matched_row,spreadsheet):
        ''' Extract information associated with the query inputted in get_row() 
            from the spreadsheet and assign it to the matching keys items in the
            header_contents dict
        '''
        try:
            for i in range(1,100):
                column = spreadsheet.cell(row=self.row_header,column=i).value
                get_info = spreadsheet.cell(row=matched_row,column=i).value
                
                if column in ExtractInfo.header_contents:
                    ExtractInfo.header_contents[column]=get_info
                    if ExtractInfo.header_contents.get(column) in (None,0):
                        ExtractInfo.header_contents[column] = "-"

        except openpyxl.utils.exceptions.InsufficientCoordinatesException:
            pass

        return ExtractInfo.header_contents


if __name__ == '__main__':
    main()
