import openpyxl
import sys
from openpyxl.styles import Border, Alignment, Font, Side
from openpyxl.styles.colors import BLACK

# alter formating if it is windoself.ws

def create_template(output_path=""):
    ''' More
    '''
    template = CreatTemplate(output_path)
    template.merge_cells()
    template.border_alignment()
    template.stylise_cells()
    template.write_to_cells()
    

class CreateTemplate(self, output_path): 
    ''' Create a template report workbook
    '''

     # create worksheet
    wb = openpyxl.Workbook()
    ws = self.wb.active
    ws.title = " "

    # style variables
    center = Alignment(horizontal="center")
    center_v = Alignment(vertical="center")
    left = Alignment(horizontal="left")
    font = Font(size=14, color=BLACK)
    small_font = Font(size=8, color=BLACK)
    thin_border = Border(left=Side(border_style='thin'),
                         right=Side(border_style='thin'), 
                         top=Side(border_style='thin'), 
                         bottom=Side(border_style='thin'))

    # variable defined dependent upon system used.
    if sys.platform == "win32":
        end_title_cell = "D"
        end_entry_cell = "H"
        column = range(2,9)
        end_border = 9

    elif sys.platform in ("cygwin","linux","linux2"):
        end_title_cell = "E"
        end_entry_cell = "J"
        column = range(2,11)
        end_border = 11

    else:
        print("Unknown system platform. The system platform is "+sys.platform)


    def __init__(CreateTemplate,output_path):
        CreateTemplate.output_path = output_path

        

    def merge_cells(CreateTemplate):
        ''' Merge cells in the template
        '''
        CreateTemplate.ws.merge_cells('B1:'+CreateTemplate.end_entry_cell+'1')
        CreateTemplate.ws.merge_cells('F4:'+CreateTemplate.end_entry_cell+'4')
        CreateTemplate.ws.merge_cells('F16:'+CreateTemplate.end_entry_cell+'20')
        CreateTemplate.ws.merge_cells('B47:'+CreateTemplate.end_entry_cell+'48')
        CreateTemplate.ws.merge_cells('B16:'+CreateTemplate.end_title_cell+'20')
        CreateTemplate.ws.merge_cells('B21:'+CreateTemplate.end_title_cell+'21')
        CreateTemplate.ws.merge_cells('B33:'+CreateTemplate.end_title_cell+'33')

        for num in range(4, 16):
            if num != 9 and num != 10:
                cell_range = "".join(("B", str(num), ":", CreateTemplate.end_title_cell, str(num)))
                CreateTemplate.ws.merge_cells(cell_range)   

        CreateTemplate.ws.merge_cells('B9:'+CreateTemplate.end_title_cell+'10')

        for n in range(7, 16):
            cell_range = "".join(("F", str(n), ":", CreateTemplate.end_entry_cell, str(n)))
            CreateTemplate.ws.merge_cells(cell_range)


    def border_alignment(CreateTemplate):
        ''' Apply borders and alignments to selected cells
        '''
        # apply border and alignment
        row = range(7,12)
        for x in row:
            CreateTemplate.ws.cell(row=x, column=CreateTemplate.end_border).border = Border(CreateTemplate.left=Side
                                                              (border_style='thin'))
            for y in column:
                CreateTemplate.ws.cell(row=x, column=y).border = CreateTemplate.thin_border
  

        for i in range(6,CreateTemplate.end_border):
            CreateTemplate.ws.cell(row=10, column=i).border = Border(top=Side(border_style=None))
            CreateTemplate.ws.cell(row=9, column=i).border = Border(bottom=Side(border_style=None))

        CreateTemplate.ws.cell(row=10,column=CreateTemplate.end_border-1).border = Border(right=Side(border_style="thin"))
        CreateTemplate.ws.cell(row=9,column=CreateTemplate.end_border-1).border = Border(right=Side(border_style="thin"))



    def stylise_cells(CreateTemplate):
        ''' Stylise selected cells
        '''
        for i in range(6,CreateTemplate.end_border-1):
            CreateTemplate.ws.cell(row=4, column=i).border = CreateTemplate.thin_border
        
        # stylise entry cells and their title cells
        for i in range(7,17):
            cell = CreateTemplate.ws.cell(row=i,column=6)
            cell.alignment = Alignment(horizontal="CreateTemplate.center", vertical="CreateTemplate.center")
            cell.CreateTemplate.font = Font(size=10, color=BLACK)
            title_cell = CreateTemplate.ws.cell(row=i, column=2)
            title_cell.CreateTemplate.font = Font(size=14, color=BLACK)
        
        for i in range(7,12):
            CreateTemplate.ws.cell(row=x, column=2).alignment = CreateTemplate.CreateTemplate.center_v
        
        comment_cell = CreateTemplate.ws.cell(row=16, column=6)
        comment_cell.CreateTemplate.font = Font(size=6, color=BLACK)
        comment_cell.style.alignment.wrap_text = True
        CreateTemplate.ws["F4"].alignment = Alignment(horizontal="CreateTemplate.center", vertical="CreateTemplate.center")
        CreateTemplate.ws["F4"].CreateTemplate.font = Font(size=10, color=BLACK)

        # change column width
        CreateTemplate.ws.column_dimensions["E"].width = 7
        CreateTemplate.ws.row_dimensions[2].height = 1
        CreateTemplate.ws.row_dimensions[3].height = 7
        CreateTemplate.ws.row_dimensions[5].height = 5

    def write_to_cells(CreateTemplate):
        ''' Write subtitles
        '''
        # write to cells
        Title = CreateTemplate.ws['B1']
        CreateTemplate.ws['B1'] = "Variant Confirmation Report"
        Title.CreateTemplate.font = Font(size=20,underline='single')
        Title.alignment = CreateTemplate.center
        CreateTemplate.ws.row_dimensions[1].height = 25

        ID = CreateTemplate.ws['B4']
        CreateTemplate.ws['B4'] = "Sample ID:"
        ID.CreateTemplate.font = CreateTemplate.font 
        ID.alignment = CreateTemplate.left
        CreateTemplate.ws.row_dimensions[4].height = 18

        Describe = CreateTemplate.ws['B6']
        CreateTemplate.ws['B6'] = "Description:"
        Describe.CreateTemplate.font = CreateTemplate.font
        Describe.alignment = CreateTemplate.left
        CreateTemplate.ws.row_dimensions[6].height = 18

        Fluidigm = CreateTemplate.ws['B21']
        CreateTemplate.ws['B21'] = "Validation:"
        Fluidigm.CreateTemplate.font = CreateTemplate.font
        Fluidigm.alignment = CreateTemplate.left
        CreateTemplate.ws.row_dimensions[21].height = 18

        Validation = CreateTemplate.ws['B34']
        CreateTemplate.ws['B34'] = "Fluidigm Image:"
        Validation.CreateTemplate.font = CreateTemplate.font
        Validation.alignment = CreateTemplate.left
        CreateTemplate.ws.row_dimensions[34].height = 18

        CreateTemplate.ws['B7'] = "Gene"
        CreateTemplate.ws['B8'] = "Exon"
        CreateTemplate.ws['B9'] = "Sequence Variant"
        CreateTemplate.ws['B11'] = "Variant Location (GRCh37)"
        # CreateTemplate.ws.row_dimensions[11].height = 30
        CreateTemplate.ws['B12'] = "Allele Balance"
        CreateTemplate.ws['B13'] = "Allele Depth (REF,ALT)"
        # CreateTemplate.ws.row_dimensions[13].height = 30
        CreateTemplate.ws['B14'] = "Allele Frequency (ESP,ExAC,dbSNP)"
        # CreateTemplate.ws.row_dimensions[14].height = 30
        CreateTemplate.ws['B15'] = "Variant Found"
        CreateTemplate.ws['B16'] = "Comment"

        Footer = CreateTemplate.ws['B47']
        self.ws['B47'] = ("The following information is for research purpose only. Any decisions made on the information should be made by an appropriate\nresponsible clinician who may require further confirmation within a clinical laboratory.")
        Footer.self.font = Font(size=8)
        Footer.alignment = self.center
        Footer.style.alignment.wrap_text = True


        self.wb.save(output_path+"test_template.xlsx")


create_template()
