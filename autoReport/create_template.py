import openpyxl
import sys
from openpyxl.styles import Border, Alignment, Font, Side
from openpyxl.styles.colors import BLACK


def create_template(output_path=""):
    ''' Create a template report workbook
    '''
    # variable defined dependent upon system used.
    if sys.platform == "win32":
        end_title_cell = "D"
        start_entry_cell = "E"
        start_entry_cell_num = 5
        end_entry_cell = "H"
        column = range(2,9)
        end_border = 9

    elif sys.platform in ("cygwin","linux","linux2"):
        end_title_cell = "E"
        start_entry_cell = "F"
        start_entry_cell_num = 6
        end_entry_cell = "J"
        column = range(2,11)
        end_border = 11

    else:
        print("Unknown system platform. The system platform is "+sys.platform)


    # create worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = " "


    # merge cells
    ws.merge_cells('B1:'+end_entry_cell+'1')
    ws.merge_cells(start_entry_cell+'4:'+end_entry_cell+'4')
    ws.merge_cells(start_entry_cell+'16:'+end_entry_cell+'20')
    ws.merge_cells('B47:'+end_entry_cell+'48')
    ws.merge_cells('B16:'+end_title_cell+'20')
    ws.merge_cells('B21:'+end_title_cell+'21')
    ws.merge_cells('B33:'+end_title_cell+'33')



    for num in range(4, 16):
        if num != 9 and num != 10:
            cell_range = "".join(("B", str(num), ":", end_title_cell, str(num)))
            ws.merge_cells(cell_range)   

    ws.merge_cells('B9:'+end_title_cell+'10')

    for n in range(7, 16):
        cell_range = "".join((start_entry_cell, str(n), ":", end_entry_cell, str(n)))
        ws.merge_cells(cell_range)


    # style variables
    center = Alignment(horizontal="center")
    center_v = Alignment(vertical="center")
    left = Alignment(horizontal="left")
    font = Font(size=14, color=BLACK)
    small_font = Font(size=8, color=BLACK)
    thin_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                         top=Side(border_style='thin'), bottom=Side(border_style='thin'))


    # apply border and alignment
    row = range(7,21)
    for i in row:
        ws.cell(row=i, column=end_border).border = Border(left=Side(border_style='thin'))
        ws.cell(row=i, column=2).alignment = center_v
        
        for x in column:
            ws.cell(row=i, column=x).border = thin_border
       
    
    for i in range(start_entry_cell_num,end_border):
        ws.cell(row=10, column=i).border = Border(top=Side(border_style=None))
        ws.cell(row=9, column=i).border = Border(bottom=Side(border_style=None))

    ws.cell(row=10,column=end_border-1).border = Border(right=Side(border_style="thin"))
    ws.cell(row=9,column=end_border-1).border = Border(right=Side(border_style="thin"))


    for i in range(start_entry_cell_num,end_border):
        ws.cell(row=4, column=i).border = thin_border




    # stylise entry cells and their title cells
    for i in range(7,17):
        cell = ws.cell(row=i,column=start_entry_cell_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=10, color=BLACK)
        title_cell = ws.cell(row=i, column=2)
        title_cell.font = Font(size=14, color=BLACK)

    
    comment_cell = ws.cell(row=16, column=start_entry_cell_num)
    comment_cell.font = Font(size=6, color=BLACK)
    comment_cell.style.alignment.wrap_text = True
    ws["F4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F4"].font = Font(size=10, color=BLACK)

    # change column width
    ws.column_dimensions["E"].width = 7
    ws.row_dimensions[2].height = 1
    ws.row_dimensions[3].height = 7
    ws.row_dimensions[5].height = 5

    # write to cells
    Title = ws['B1']
    ws['B1'] = "Variant Confirmation Report"
    Title.font = Font(size=20,underline='single')
    Title.alignment = center
    ws.row_dimensions[1].height = 25

    ID = ws['B4']
    ws['B4'] = "Sample ID:"
    ID.font = font 
    ID.alignment = left
    ws.row_dimensions[4].height = 18

    Describe = ws['B6']
    ws['B6'] = "Description:"
    Describe.font = font
    Describe.alignment = left
    ws.row_dimensions[6].height = 18

    Fluidigm = ws['B21']
    ws['B21'] = "Validation:"
    Fluidigm.font = font
    Fluidigm.alignment = left
    ws.row_dimensions[21].height = 18

    Validation = ws['B34']
    ws['B34'] = "Fluidigm Image:"
    Validation.font = font
    Validation.alignment = left
    ws.row_dimensions[34].height = 18

    ws['B7'] = "Gene"
    ws['B8'] = "Exon"
    ws['B9'] = "Sequence Variant"
    ws['B11'] = "Variant Location (GRCh37)"
    # ws.row_dimensions[11].height = 30
    ws['B12'] = "Allele Balance"
    ws['B13'] = "Allele Depth (REF,ALT)"
    # ws.row_dimensions[13].height = 30
    ws['B14'] = "Allele Frequency (ESP,ExAC,dbSNP)"
    # ws.row_dimensions[14].height = 30
    ws['B15'] = "Variant Found"
    ws['B16'] = "Comment"



    Footer = ws['B47']
    ws['B47'] = ("The following information is for research purpose only. Any decisions made on the information should be made by an appropriate\nresponsible clinician who may require further confirmation within a clinical laboratory.")
    Footer.font = Font(size=8)
    Footer.alignment = center
    Footer.style.alignment.wrap_text = True


    wb.save(output_path+"test_template.xlsx")

